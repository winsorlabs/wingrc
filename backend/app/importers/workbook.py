"""Importer for the Authorized-Entities workbook.

Parses the four tabs into `CanonicalEntity` records, inferring CMMC scope
category where the source provides it. Raw column values are preserved under
their original headers so the renderer can reproduce the list faithfully.

The same shape of importer will exist for CSV, Liongard and Datto RMM — each
one's only job is: source rows -> List[CanonicalEntity]. Reconciliation and
rendering are shared downstream.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from ..catalog import (
    AUTHORIZED_DEVICES,
    AUTHORIZED_PROCESSES,
    AUTHORIZED_USERS,
    EXTERNAL_SERVICES,
    ListView,
)
from ..domain import (
    CanonicalEntity,
    EntityStatus,
    EntityType,
    ScopeCategory,
    Source,
)

_PLACEHOLDER_TOKENS = ("[placeholder]",)
_CATEGORY_LOOKUP = {c.value.lower(): c for c in ScopeCategory}


def _is_placeholder(value: Any) -> bool:
    return isinstance(value, str) and any(
        t in value.lower() for t in _PLACEHOLDER_TOKENS
    )


def _infer_category(attributes: dict[str, Any]) -> ScopeCategory | None:
    """Look for a known CMMC category token in the columns that tend to hold it."""
    for col in ("Asset Type", "Owner / Primary User"):
        raw = attributes.get(col)
        if isinstance(raw, str):
            cat = _CATEGORY_LOOKUP.get(raw.strip().lower())
            if cat:
                return cat
    return None


def _header_row_index(rows: list[tuple], view: ListView) -> int:
    """Find the row whose cells match the view's first column header."""
    first_header = view.columns[0][1]
    for i, row in enumerate(rows):
        if row and any(
            isinstance(c, str) and c.strip() == first_header for c in row
        ):
            return i
    raise ValueError(f"Header row not found for sheet {view.sheet_title!r}")


def _natural_key(view: ListView, attributes: dict[str, Any]) -> str:
    if view is AUTHORIZED_USERS:
        return f"{attributes.get('First Name', '')} {attributes.get('Last Name', '')}".strip()
    if view is AUTHORIZED_PROCESSES:
        return str(attributes.get("Process Name", "")).strip()
    if view is AUTHORIZED_DEVICES:
        serial = attributes.get("Serial # or Asset Tag")
        return str(serial or attributes.get("Name", "")).strip()
    if view is EXTERNAL_SERVICES:
        return str(attributes.get("Name", "")).strip()
    return str(next(iter(attributes.values()), "")).strip()


def _entity_type(view: ListView) -> EntityType:
    return view.entity_type


def parse_workbook(path: str | Path) -> list[CanonicalEntity]:
    """Parse every supported tab of the workbook into canonical entities."""
    wb = openpyxl.load_workbook(path, data_only=True)
    views_by_sheet = {
        v.sheet_title: v
        for v in (
            AUTHORIZED_USERS,
            AUTHORIZED_PROCESSES,
            AUTHORIZED_DEVICES,
            EXTERNAL_SERVICES,
        )
    }
    entities: list[CanonicalEntity] = []

    for sheet_title, view in views_by_sheet.items():
        if sheet_title not in wb.sheetnames:
            continue
        ws = wb[sheet_title]
        rows = list(ws.iter_rows(values_only=True))
        try:
            header_idx = _header_row_index(rows, view)
        except ValueError:
            continue

        header_row = rows[header_idx]
        # Map each source column position to the header text we care about.
        positions = {
            j: str(cell).strip()
            for j, cell in enumerate(header_row)
            if isinstance(cell, str) and cell.strip()
        }

        for row in rows[header_idx + 1 :]:
            if not row or all(c is None for c in row):
                continue
            attributes = {
                positions[j]: row[j]
                for j in positions
                if j < len(row) and row[j] is not None
            }
            if not attributes:
                continue
            # Skip illustrative placeholder rows.
            if any(_is_placeholder(v) for v in attributes.values()):
                continue

            natural_key = _natural_key(view, attributes)
            if not natural_key:
                continue

            category = _infer_category(attributes)
            decommissioned = bool(attributes.get("Decommissioned Date"))

            entities.append(
                CanonicalEntity(
                    entity_type=_entity_type(view),
                    natural_key=natural_key,
                    attributes=attributes,
                    scope_category=category,
                    status=(
                        EntityStatus.DECOMMISSIONED
                        if decommissioned
                        else EntityStatus.ACTIVE
                    ),
                    in_boundary=True,
                    source=Source.WORKBOOK,
                    source_ref=str(Path(path).name),
                )
            )

    return entities
