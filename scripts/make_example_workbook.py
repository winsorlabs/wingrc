"""Generate a sanitized example Authorized-Entities workbook for the repo.

Fictional data only — no customer environment details. This doubles as the
import fixture for tests and the 'try it now' sample in the quickstart.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font

SHEETS = {
    "3.1.1a Authorized Users": {
        "title": "Authorized Users - Active Directory",
        "control": "AC.L2-3.1.1",
        "headers": [
            "First Name", "Last Name", "Data Intake Admin", "Remote Access",
            "Requested By/Responsible Party", "Start Date", "End Date",
        ],
        "rows": [
            ["Ada", "Lovelace", "Yes", "Yes", "Security Officer", "2025-01-06", ""],
            ["Grace", "Hopper", "No", "Yes", "Security Officer", "2025-02-03", ""],
            ["Alan", "Turing", "No", "No", "IT Manager", "2025-03-17", ""],
        ],
    },
    "3.1.1b Auth Processes": {
        "title": "Processes Acting on Behalf of Authorized Users",
        "control": "AC.L2-3.1.1",
        "headers": ["Process Name", "Running On", "Associated Account", "Description / Purpose"],
        "rows": [
            ["BackupAgentSvc", "SRV-FILE01", "svc-backup", "Nightly CUI backup job"],
            ["PatchOrchestrator", "SRV-RMM01", "svc-rmm", "Scheduled patch deployment"],
        ],
    },
    "3.1.1c Authorized Devices": {
        "title": "Authorized Devices",
        "control": "AC.L2-3.1.1",
        "headers": [
            "Name", "Owner / Primary User", "Make", "Model", "Serial # or Asset Tag",
            "Mac Address", "OS", "BIOS FW Ver", "Location", "Asset Type",
            "In Service Date", "Decommissioned Date", "FenixPyre Installed",
            "DUO Installed", "Senteon Installed", "RoboShadow Installed", "Heimdal Installed",
        ],
        "rows": [
            ["WS-0001", "CUI Asset", "Dell", "Latitude 7440", "ASSET-0001",
             "00:11:22:33:44:55", "Windows 11 Pro 24H2", "1.20.0", "HQ - Suite 200",
             "CUI Asset", "2025-01-10", "", "Y", "Y", "Y", "Y", "Y"],
            ["SRV-FILE01", "CUI Asset", "Dell", "PowerEdge R660", "ASSET-0002",
             "00:11:22:33:44:66", "Windows Server 2022", "2.10.1", "HQ - Server Room",
             "CUI Asset", "2025-01-10", "", "Y", "N", "Y", "N", "Y"],
            ["FW-EDGE01", "SPA", "Fortinet", "FortiGate 70G", "ASSET-0003",
             "00:11:22:33:44:77", "FortiOS 7.6.7", "", "HQ - Server Room",
             "SPA", "2025-01-10", "", "N", "N", "N", "Y", "N"],
        ],
    },
    "External Services": {
        "title": "External / Cloud Services",
        "control": "AC.L2-3.1.1 / 3.1.20",
        "headers": ["Name", "Provider", "Asset Type"],
        "rows": [
            ["Heimdal", "ESP", "SPA"],
            ["Datto RMM", "ESP", "SPA"],
            ["Liongard", "ESP", "SPA"],
            ["Microsoft 365 GCC High", "CSP", "CRMA"],
            ["DUO", "ESP", "SPA"],
        ],
    },
}


def build(out_path: Path) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_title, spec in SHEETS.items():
        ws = wb.create_sheet(title=sheet_title[:31])
        ws["C1"] = "Authorized Entities"
        ws["C2"] = f"CMMC Practice / NIST 800-171: {spec['control']}"
        ws["A8"] = spec["title"]
        ws["A8"].font = Font(bold=True)
        for col, header in enumerate(spec["headers"], start=1):
            ws.cell(row=10, column=col, value=header).font = Font(bold=True)
        for r, row in enumerate(spec["rows"], start=11):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    p = build(Path(__file__).resolve().parents[1] / "samples" / "authorized-entities.example.xlsx")
    print(f"Wrote {p}")
